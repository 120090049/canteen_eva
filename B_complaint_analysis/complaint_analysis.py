import pandas as pd
import openpyxl
import os

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)

file_path = os.path.join(parent_dir, 'data', 'complaint_score.xlsx')

def get_keys(my_dict, value):
    keys = []
    for k, v in my_dict.items():
        if v == value:
            keys.append(k)
    if len(keys) == 0:
        return None
    return keys


def complaint_analysis(stall_dict, canteen_dict):
    table = openpyxl.load_workbook(file_path).active
    stall_score_list = [5] * 45
    first_row = True
    for row_index in range(1,table.max_row+1):  
        if (first_row):
            first_row = False
            continue
        else:
            canteen = table.cell(row_index, 1).value
            stall = table.cell(row_index, 2).value
            score = table.cell(row_index, 5).value
            
            if (stall == "未知"): # all stalls in a certain canteens should be deducted
                indexs = get_keys(canteen_dict, canteen)
                for index in indexs:
                    stall_score_list[index-1] -= (1-score)/len(indexs)
                    
            else:   # deduct a certain stall
                intersection = (set(get_keys(stall_dict, stall)) & set(get_keys(canteen_dict, canteen)))
                index = intersection.pop()
                stall_score_list[index-1] -= (1-score)
    return stall_score_list
 
