import openpyxl
import numpy as np
# from numpy.linalg import *
from numpy.linalg import eig
import os


# read the excel and form a table
# Define variable to load the dataframe
def cuisine_score():
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # 构造文件的相对路径
    file_path = os.path.join(current_dir, "matrixH.xlsx")

    dataframe = openpyxl.load_workbook(file_path)

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # store the raw data
    table = []
    row_num = dataframe1.max_row
    col_num = dataframe1.max_column

    # read Excel to table
    for i in range(0, row_num):
        row = []
        for j in dataframe1.iter_cols(1, dataframe1.max_column):
            row.append(j[i].value)
        table.append(row)
        del row

    # use table to form H matrix
    H = []

    for i in range(row_num):
        H_row = []
        total = table[i][-1]
        # print(total)
        for j in range(col_num-1):
            H_row.append(0.9*(table[i][j]/total))
        H.append(H_row)
        del H_row

    I = [[1/450]*45]*45
    I = np.array(I)

    # print(I)
    H_hat = H + I

    # pi
    pi = [0.5, 0, 0, 0, 0,  0, 0, 0, 0, 0,
          0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
          0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
          0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
          0, 0, 0, 0, 0.5]

    # iteration to gain the convergent result--pi
    for i in range(100):
        pi = np.dot(pi, H_hat)
    # print('k= ', 100, ' :', pi)

    # check the value
    sum = 0
    for i in range(45):
        sum = sum+pi[i]
        # element = pi[i]
        # pi[i] = 1-element

    # sorted list
    new_list = np.argsort(pi).tolist()  # now, the stall index from 0-44

    # let new_list contains stall index from 1 to 45
    # for i in range(45):
    # element = new_list[i]
    # new_list[i] = element+1

    # print("sum = ", sum)
    # print(pi)
    # print(new_list)

    # to record the borda count same rank has the same count, therefore, no 0 count
    # the borda count list[0] is the borda count of stall_1
    borda_list = [0]*45

    last_position = 0
    last_borda = 44
    # print("index = ", new_list[0]+1, "borda count = ", 44)

    borda_list[new_list[0]] = 44
    for k in range(1, 45):
        if pi[k] > pi[new_list[last_position]]:
            # print("index = ", new_list[k]+1, "borda count = ", 44-k)
            borda_list[new_list[k]] = 44-k
            last_position = k
            last_borda = 44-k
        else:
            # print("index = ", new_list[k]+1, "borda count = ", last_borda)
            borda_list[new_list[k]] = last_borda

    # print(pi)
    final_result = []
    for i in range(len(pi)):
        final_result.append(1 / pi[i])
        # print(i, "hhh", 1/pi[i])
    print(final_result)

    # return new_list
    return final_result

# *************************** PS *************************
# 直接用所有的来形成matrix，现在结果中有0存在，这不符合之前的设定
# 有尝试过加上thelta，结果比较接近，差异较小


if __name__ == "__main__":
    similarity_list = cuisine_score()
    # print(similarity_list)
