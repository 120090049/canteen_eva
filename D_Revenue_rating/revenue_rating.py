import openpyxl
import numpy as np
from numpy.linalg import *
import os

def revenue_rating():

    # Define variable to load the dataframe
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # 构造文件的相对路径
    file_path = os.path.join(current_dir, "readTest.xlsx")
    
    dataframe = openpyxl.load_workbook(file_path)

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # store the raw data
    table = []
    row_num = dataframe1.max_row
    col_num = dataframe1.max_column

    # average of revenue
    r = 117450 

    # A=x*(7+41) B=(7+41)*1 C=x*1
    A_row = [0]*46
    A = []
    C = []

    # record the missed data position
    needPredict = []

    # read Excel to table
    for i in range(0, row_num):
        row = []
        for j in dataframe1.iter_cols(1, dataframe1.max_column):
            row.append(j[i].value)
        table.append(row)
        del row

    # from matrix table into A and c
    count = 0  # to record the number of known data
    for i in range(row_num):
        for j in range(col_num):
            # 对所有的点都要查看，所以是在两个for的下面
            if table[i][j] != None:
                # 对A
                A_row[i] = 1
                A_row[41+j] = 1
                count = count+1
                # print(A_row)
                A.append(A_row)
                A_row = [0]*46  # recover the row vector

                # 对c
                C.append([table[i][j]-r]) #let vector C to be colume vector

            else:
                need = [i, 41+j] #record the missed postion
                needPredict.append(need)


    # solve the optimal b*
    A = np.array(A)
    C = np.array(C)
    # print(C)
    AT = A.T
    # print(AT)
    ATA = np.dot(AT, A)
    ATC = np.dot(AT, C)

    B = solve(ATA, ATC)

    # print out the predicted data 
    for cell in range(29):
        b1 = needPredict[cell][0]
        b2 = needPredict[cell][1]
        result = B[b1][0] + B[b2][0] + r
        # print("b1 = ", b1, "b2 = ", b2, "result = ", result)

    # CANNOT WRITE into this sheet, so I copy it into Excel manually

    # read another sheet to print out the rank list
    # Define variable to load the dataframe

    # 构造文件的相对路径
    file_path = os.path.join(current_dir, 'Stall Revenue.xlsx')
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Netflix(v1)"]

    rank = ws['M']
    rank_list = []
    for x in range(1,len(rank)): 
        # print(rank[x].value) 
        rank_list.append(rank[x].value)
    

    # new_list = np.argsort(rank_list).tolist()
    # sorted_indexes = new_list[:-1]
    # # print(new_list)
    
    # index_score = [0 for i in range(len(sorted_indexes))]
    # for i in range(len(sorted_indexes)):
    #     score = len(sorted_indexes) - i -1
    #     index_score[sorted_indexes[i]-1] = score
    
    return rank_list

if __name__ == "__main__":
    revenue_list = revenue_rating()
    print(revenue_list)