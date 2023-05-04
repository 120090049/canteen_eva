from A_bayesian_ranking.bayesian_ranking import bayesian_ranking
from B_complaint_analysis.complaint_analysis import complaint_analysis
from C_Modified_Page_Ranking.geo_score import geo_score
from C_Modified_Page_Ranking.cuisine_score import cuisine_score
from D_Revenue_rating.revenue_rating import revenue_rating
from borda import borda, infer

import os
import openpyxl
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import numpy as np

remove_percent = 0.1
calculate = False

def ori2borda(original_score):
    sorted_score = sorted(original_score)  # 按升序排序 [2, 3, 3, 4, 5, 6, 6, 6]
    result = [0 for i in range(len(original_score))]

    for i in range(len(original_score)):
        result[i] = sorted_score.index(original_score[i]) 
    # print(result) # [4, 0, 1, 1, 3, 5, 5, 5]  8个
    count_dict = {}
    for num in result:
        if num in count_dict:
            count_dict[num] += 1
        else:
            count_dict[num] = 1

    index_repeated = [key for key, value in count_dict.items() if value != 1] 

    new_dict = {} # {1: 2, 5: 3} ==> {1:1.5, 5:6}
    for i in index_repeated:
        # 1 有 2 个 == i 有 times 个   1+2+3
        times = count_dict[i] 
        value = i + 0.5*(times-1)
        new_dict[i] = value        
    # number_index_repeated = 

    for i in range(len(result)):
        if result[i] in new_dict:
            result[i] = new_dict[result[i]]

            
    return result



if __name__ == "__main__":
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, 'data', 'stall_canteen_campus.xlsx')

    # 读取Excel文件
    table = openpyxl.load_workbook(file_path)
    table_sheet = table.active

    first_row = True
    
    stall_dict = {} # {stall_name: index}
    canteen_dict = {} #  {stall_name: canteen}
    campus_dict = {}
    for row_index in range(1,table_sheet.max_row+1):
        # 读取第一列和第二列的值
        canteen = table_sheet.cell(row_index, 1).value
        stall = table_sheet.cell(row_index, 2).value
        campus = table_sheet.cell(row_index, 3).value

        if (first_row):
            first_row = False
            continue
        else:
            stall_dict[int(row_index)-1] = stall
            canteen_dict[int(row_index)-1] = canteen
            campus_dict[canteen] = campus

    list_A = bayesian_ranking()
    list_B = complaint_analysis(stall_dict, canteen_dict)
    list_C = geo_score(stall_dict, canteen_dict)
    list_D = cuisine_score()
    list_E = revenue_rating()
    # print(list_A)
    # print(list_B)
    # print(list_C)
    # print(list_D)
    # print(list_E)
    lists = [ori2borda(list_A), ori2borda(list_B), ori2borda(list_C), ori2borda(list_D), ori2borda(list_E)]
    if (calculate):
        final_result = borda(remove_percent, lists)
    else:
        result = [9, 57, 16, 3, 15, 65.923273657289, [30, 29, 44, 31]]
        w1, w2, w3, w4, w5 = result[0], result[1], result[2], result[3], result[4]
        print("w1 = %d, w2 = %d, w3 = %d, w4 = %d, w5 = %d and corresponding cost is %d" % (w1, w2, w3, w4, w5, result[5]))
        print()
        print("The stall need to be removed:")
        for index in result[6]:
            for stall_index, stall_name in stall_dict.items():
                if int(stall_index) == index:
                    # found the stall with the desired index
                    print(str(index) +": " + stall_name)
                    break
    remove_num = int( len(canteen_dict) * remove_percent )
    print(remove_num)
    list, out = infer(lists, [w1, w2, w3, w4, w5], remove_num)
    print(list, out)
    # # Create the data for the plot
    # sum = 100 - w4 - w5 
    # x = np.arange(0, (sum+1))
    # y = np.arange(0, (sum+1))
    # x = np.meshgrid(x, y)
    
    
    # remove_num = int( len(canteen_dict) * remove_percent )
    
    # x, y = np.meshgrid(x, y)
    # res = np.zeros_like(x)

    # # Evaluate the function at each point in the grid
    # for i in range(sum):
    #     for j in range(sum):
    #         print(x[i,j], y[i,j])
    #         if x[i,j] + y[i,j] <= sum:
    #             # print( [x[i,j], y[i,j],(sum-x[i,j]-y[i,j]),w4,w5])
    #             _, out = infer(lists, [x[i,j], y[i,j],(sum-x[i,j]-y[i,j]),w4,w5], remove_num)
    #             res[i,j] = out
    #             print(out)
                

    # # Plot the surface
    # fig = plt.figure()
    # ax = fig.add_subplot(111, projection='3d')
    # ax.plot_surface(x, y, res)

    # # Set the labels and title
    # ax.set_xlabel('x')
    # ax.set_ylabel('y')
    # ax.set_zlabel('z')
    # ax.set_title('3D plot of z = f(x,y)')

    # # Show the plot
    # plt.show()



