import pandas as pd
import os
import openpyxl


if __name__ == "__main__":
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, 'stall_rating.xlsx')

    # 读取Excel文件
    table = openpyxl.load_workbook(file_path)
    table_sheet = table.active

    first_row = True
    stall_dict = {}
    for row_index in range(1,table_sheet.max_row+1):
        # 读取第一列和第二列的值
        canteen = table_sheet.cell(row_index, 1).value
        stall = table_sheet.cell(row_index, 2).value

        if (first_row):
            first_row = False
            continue
        else:
            stall_dict[row_index-1] = [stall, canteen]
    
    print(stall_dict)
    
    # final_list = bayesian_ranking()
    # print(final_list)
